import json
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import datetime
from apps.accounts.decorators import owner_required
from .services import get_all_report_data, get_export_orders_data, generate_feedback_wordcloud
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

@login_required
@owner_required
def dashboard(request):
    today = timezone.now().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date = today.replace(day=1)
            end_date = today
    else:
        start_date = today.replace(day=1)
        end_date = today

    # Ambil semua data dari services
    data = get_all_report_data(start_date, end_date)

    # Konversi list ke JSON string agar aman di JavaScript
    data['daily_trend']['dates'] = json.dumps(data['daily_trend']['dates'])
    data['daily_trend']['revenues'] = json.dumps(data['daily_trend']['revenues'])
    data['top_packages']['labels'] = json.dumps(data['top_packages']['labels'])
    data['top_packages']['data'] = json.dumps(data['top_packages']['data'])

    # Hitung persentase untuk progress bar
    total_fb = data['sentiment']['positive'] + data['sentiment']['neutral'] + data['sentiment']['negative']
    data['sentiment']['positive_percentage'] = (data['sentiment']['positive'] / total_fb * 100) if total_fb > 0 else 0

    wordcloud_img = generate_feedback_wordcloud(start_date, end_date)
    data['wordcloud_img'] = wordcloud_img

    data['start_date'] = start_date.isoformat()
    data['end_date'] = end_date.isoformat()

    return render(request, 'reports/dashboard.html', data)


@login_required
@owner_required
def export_orders_excel(request):
    today = timezone.now().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date = today.replace(day=1)
            end_date = today
    else:
        start_date = today.replace(day=1)
        end_date = today

    orders_data = get_export_orders_data(start_date, end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Order"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C8251F", end_color="C8251F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Kode Nota", "Customer", "Nomor HP", "Paket", "Berat (kg)",
        "Total Harga", "Progress", "Pembayaran", "Tanggal Order", "Tanggal Update",
        "Waktu Dicuci", "Waktu Dikeringkan", "Waktu Disetrika", "Waktu Selesai", "Waktu Diambil"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_idx, order in enumerate(orders_data, start=2):
        ws.cell(row=row_idx, column=1, value=order['kode_nota'])
        ws.cell(row=row_idx, column=2, value=order['nama_customer'])
        ws.cell(row=row_idx, column=3, value=order['nomor_hp'])
        ws.cell(row=row_idx, column=4, value=order['paket'])
        ws.cell(row=row_idx, column=5, value=order['berat'])
        ws.cell(row=row_idx, column=6, value=order['total_harga'])
        ws.cell(row=row_idx, column=7, value=order['progress_status'])
        ws.cell(row=row_idx, column=8, value=order['payment_status'])
        ws.cell(row=row_idx, column=9, value=order['tanggal_order'])
        ws.cell(row=row_idx, column=10, value=order['tanggal_update'])
        ws.cell(row=row_idx, column=11, value=order['waktu_dicuci'])
        ws.cell(row=row_idx, column=12, value=order['waktu_dikeringkan'])
        ws.cell(row=row_idx, column=13, value=order['waktu_disetrika'])
        ws.cell(row=row_idx, column=14, value=order['waktu_selesai'])
        ws.cell(row=row_idx, column=15, value=order['waktu_diambil'])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        adjusted = (max_len + 2) * 1.2
        ws.column_dimensions[col_letter].width = min(adjusted, 30)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="laporan_order_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response